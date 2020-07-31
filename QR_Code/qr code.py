import openpyxl
import qrcode

farmer_details_path='Database_Of _Farmers/Farmlanddatabase.xlsx'
wkb_obj = openpyxl.load_workbook(farmer_details_path)
sheet_obj = wkb_obj.active

rows= sheet_obj.max_row
for i in range(2,rows+1):
    myqr = qrcode.QRCode()

    myqr.add_data('Farmer Id: ' + str(sheet_obj.cell(row=i,column=2).value))
    myqr.add_data('\n')
    myqr.add_data('Name: ' + str(sheet_obj.cell(row=i,column=3).value))
    myqr.add_data('\n')
    myqr.add_data('District: ' + str(sheet_obj.cell(row=i,column=4).value))
    myqr.add_data('\n')
    myqr.add_data('Village: ' + str(sheet_obj.cell(row=i,column=5).value))
    myqr.add_data('\n')
    myqr.add_data('Land No: ' + str(sheet_obj.cell(row=i,column=6).value))
    myqr.add_data('\n')
    myqr.add_data('Land Area(in acre): ' + str(sheet_obj.cell(row=i,column=7).value))
    myqr.add_data('\n')
    myqr.add_data('Soil Type: ' + str(sheet_obj.cell(row=i,column=8).value))
    myqr.add_data('\n')
    myqr.add_data('Latitude: ' + str(sheet_obj.cell(row=i,column=9).value))
    myqr.add_data('\n')
    myqr.add_data('Longitude: ' + str(sheet_obj.cell(row=i,column=10).value))
    myqr.add_data('\n')
    myqr.add_data('Water Source: ' + str(sheet_obj.cell(row=i,column=11).value))
    myqr.make()
    img=myqr.make_image()
    img.save('QR_Code/Sample QR/'+'myqr'+str(i-1)+'.png')
